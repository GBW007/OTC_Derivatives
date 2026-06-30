"""
excel_reporter.py — Excel Report Generator
============================================
Generates a professional multi-sheet Excel workbook summarising
the entire OTC derivatives simulation.

Sheets produced:
  1. Summary Dashboard  — High-level KPIs and portfolio overview
  2. Trade Register     — All booked trades
  3. MTM Report         — Daily MTM by trade (last day)
  4. Margin Calls       — All margin calls issued
  5. Collateral         — Collateral posted and received
  6. Lifecycle Events   — Trade state changes
  7. Reconciliation     — Accuracy and breaks report

Uses openpyxl for formatting (colours, borders, bold headers).
"""

import logging
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, PieChart
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from config import EXCEL_OUTPUT_PATH

logger = logging.getLogger(__name__)


# ── Colour palette ────────────────────────────────────────────
NAVY    = "1F3864"   # Header background
GOLD    = "C9A84C"   # Accent / highlight
WHITE   = "FFFFFF"
LIGHT_BLUE = "DEEAF1"
LIGHT_GREEN = "E2EFDA"
LIGHT_RED   = "FFDEDE"
GREY    = "F2F2F2"


class ExcelReporter:
    """Generates the OTC Derivatives Excel Report."""

    def __init__(self, db_manager):
        self.db = db_manager

    def generate_report(self, output_path: str = EXCEL_OUTPUT_PATH,
                        simulation_days: int = 30) -> str:
        """
        Generate the full Excel workbook.

        Args:
            output_path:     Where to save the file
            simulation_days: How many days were simulated

        Returns:
            Path to the saved file
        """
        if not HAS_OPENPYXL:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return None

        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # Remove default empty sheet

        # Gather data
        trades        = self.db.get_all_trades()
        margin_calls  = self.db.get_margin_calls()
        settlements   = self.db.get_settlements()
        col_summary   = self.db.get_collateral_summary()
        recon_summary = self.db.get_reconciliation_summary()
        portfolio     = self.db.get_portfolio_summary()

        # ── Build each sheet ──────────────────────────────────
        self._sheet_summary(wb, trades, margin_calls, settlements,
                            col_summary, recon_summary, simulation_days)
        self._sheet_trades(wb, trades)
        self._sheet_mtm(wb)
        self._sheet_margin(wb, margin_calls)
        self._sheet_collateral(wb, col_summary)
        self._sheet_reconciliation(wb, recon_summary)

        wb.save(output_path)
        logger.info(f"[EXCEL] Report saved → {output_path}")
        return output_path

    # ── HELPER: Style a header row ─────────────────────────────

    def _style_header(self, ws, row: int, columns: list):
        """Apply navy header styling to a row."""
        for col, heading in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=col, value=heading)
            cell.font = Font(bold=True, color=WHITE, size=10)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                bottom=Side(border_style="thin", color=GOLD)
            )
        ws.row_dimensions[row].height = 20

    def _style_title(self, ws, row: int, col: int, text: str, span: int = 1):
        """Write a bold title cell."""
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, size=12, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        if span > 1:
            ws.merge_cells(
                start_row=row, start_column=col,
                end_row=row, end_column=col + span - 1
            )
        return cell

    def _zebra(self, ws, row: int, num_cols: int, idx: int):
        """Apply alternating row shading for readability."""
        colour = GREY if idx % 2 == 0 else WHITE
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=colour)

    def _autofit_columns(self, ws, min_width=8, max_width=40):
        """Set column widths based on content."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)

    # ── SHEET 1: Summary Dashboard ─────────────────────────────

    def _sheet_summary(self, wb, trades, margin_calls, settlements,
                       col_summary, recon_summary, sim_days):
        ws = wb.create_sheet("📊 Summary Dashboard")
        ws.sheet_view.showGridLines = False

        # Title
        self._style_title(ws, 1, 1, "OTC DERIVATIVES LIFECYCLE & MARGIN SIMULATION", 8)
        ws.cell(row=2, column=1,
                value=f"Report Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Simulation Period: {sim_days} Business Days")
        ws.cell(row=2, column=1).font = Font(italic=True, color="666666", size=9)
        ws.merge_cells("A2:H2")

        # ── KPI Cards ────────────────────────────────────────
        kpi_start = 4
        kpis = [
            ("Total Trades",         len(trades),                              ""),
            ("Active Trades",        sum(1 for t in trades if t["status"] == "ACTIVE"), ""),
            ("Settled Trades",       sum(1 for t in trades if t["status"] == "SETTLED"), ""),
            ("Margin Calls",         len(margin_calls),                        ""),
            ("Total IM (₹ Cr)",      round(sum(m["call_amount_inr"] for m in margin_calls
                                    if m["margin_type"]=="INITIAL_MARGIN") / 1e7, 2), "Cr"),
            ("Total VM (₹ Cr)",      round(sum(m["call_amount_inr"] for m in margin_calls
                                    if m["margin_type"]=="VARIATION_MARGIN") / 1e7, 2), "Cr"),
            ("Recon Accuracy",       f"{recon_summary.get('accuracy_pct', 0):.2f}%",   ""),
            ("Open Breaks",          recon_summary.get("BREAK", 0),           ""),
        ]

        ws.cell(row=kpi_start, column=1, value="KEY PERFORMANCE INDICATORS").font = Font(bold=True, size=11, color=NAVY)

        for i, (label, value, unit) in enumerate(kpis):
            col = (i % 4) * 2 + 1
            row = kpi_start + 2 + (i // 4) * 3

            label_cell = ws.cell(row=row, column=col, value=label)
            label_cell.font = Font(bold=True, size=9, color="444444")
            label_cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)

            val_cell = ws.cell(row=row + 1, column=col,
                               value=f"{value} {unit}".strip())
            val_cell.font = Font(bold=True, size=14, color=NAVY)
            val_cell.fill = PatternFill("solid", fgColor=WHITE)
            val_cell.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=col + 1)
            ws.merge_cells(start_row=row + 1, start_column=col,
                           end_row=row + 1, end_column=col + 1)

        # ── Trade Breakdown Table ─────────────────────────────
        tbl_row = kpi_start + 11
        ws.cell(row=tbl_row, column=1, value="TRADE BREAKDOWN BY INSTRUMENT").font = Font(bold=True, color=NAVY)
        tbl_row += 1
        self._style_header(ws, tbl_row, ["Instrument", "Count", "Notional (₹ Cr)", "Avg Notional (₹ Cr)"])
        by_inst = {}
        for t in trades:
            inst = t["instrument_type"]
            if inst not in by_inst:
                by_inst[inst] = {"count": 0, "notional": 0}
            by_inst[inst]["count"] += 1
            by_inst[inst]["notional"] += t.get("notional_inr", 0)

        for idx, (inst, data) in enumerate(by_inst.items()):
            r = tbl_row + 1 + idx
            self._zebra(ws, r, 4, idx)
            ws.cell(r, 1, inst)
            ws.cell(r, 2, data["count"])
            ws.cell(r, 3, round(data["notional"] / 1e7, 1))
            ws.cell(r, 4, round(data["notional"] / max(data["count"], 1) / 1e7, 1))

        # ── Regulatory Coverage ───────────────────────────────
        reg_row = tbl_row + len(by_inst) + 3
        ws.cell(row=reg_row, column=1, value="REGULATORY COVERAGE").font = Font(bold=True, color=NAVY)
        reg_row += 1
        self._style_header(ws, reg_row, ["Jurisdiction", "# Trades", "Framework", "Clearing"])
        indian = [t for t in trades if t.get("jurisdiction") == "INDIA"]
        foreign = [t for t in trades if t.get("jurisdiction") == "INTERNATIONAL"]
        for idx, (jur, tlist, framework, clearing) in enumerate([
            ("India",         indian,  "RBI/SEBI Master Direction", "CCIL"),
            ("International", foreign, "ISDA/Basel III / Dodd-Frank / EMIR", "LCH/CME/Bilateral"),
        ]):
            r = reg_row + 1 + idx
            self._zebra(ws, r, 4, idx)
            ws.cell(r, 1, jur)
            ws.cell(r, 2, len(tlist))
            ws.cell(r, 3, framework)
            ws.cell(r, 4, clearing)

        self._autofit_columns(ws)

    # ── SHEET 2: Trade Register ─────────────────────────────────

    def _sheet_trades(self, wb, trades):
        ws = wb.create_sheet("📝 Trade Register")
        ws.freeze_panes = "A3"
        self._style_title(ws, 1, 1, "OTC TRADE REGISTER — ALL BOOKED TRADES", 11)

        headers = [
            "Trade ID", "Instrument", "Trade Date", "Start Date",
            "Maturity Date", "Notional", "Currency", "Notional (₹ Cr)",
            "Fixed Rate", "Direction", "Counterparty", "Country",
            "Clearing", "Status", "Jurisdiction", "Regulatory Regime"
        ]
        self._style_header(ws, 2, headers)

        for i, t in enumerate(trades):
            r = 3 + i
            self._zebra(ws, r, len(headers), i)
            row_data = [
                t.get("trade_id"), t.get("instrument_type"), t.get("trade_date"),
                t.get("start_date"), t.get("maturity_date"),
                f"{t.get('notional', 0):,.0f}", t.get("currency"),
                round(t.get("notional_inr", 0) / 1e7, 2),
                f"{t.get('fixed_rate', 0)*100:.3f}%", t.get("direction"),
                t.get("counterparty_name"), t.get("counterparty_country"),
                t.get("clearing_venue"), t.get("status"),
                t.get("jurisdiction"), t.get("regulatory_regime"),
            ]
            for col, val in enumerate(row_data, start=1):
                cell = ws.cell(r, col, val)
                if t.get("status") == "SETTLED":
                    cell.font = Font(color="888888", italic=True)
                if t.get("status") == "ACTIVE":
                    ws.cell(r, 14).font = Font(color="006100", bold=True)

        self._autofit_columns(ws)

    # ── SHEET 3: MTM Report ─────────────────────────────────────

    def _sheet_mtm(self, wb):
        ws = wb.create_sheet("💹 MTM Valuations")
        ws.freeze_panes = "A3"
        self._style_title(ws, 1, 1, "DAILY MARK-TO-MARKET VALUATIONS (LATEST DAY)", 9)

        # Get the latest day's MTM from DB
        import sqlite3, os
        from config import DATABASE_PATH
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.execute("""
            SELECT m.*, t.instrument_type, t.counterparty_name, t.currency as trade_ccy
            FROM mtm_valuations m
            JOIN trades t ON m.trade_id = t.trade_id
            WHERE m.valuation_date = (SELECT MAX(valuation_date) FROM mtm_valuations)
            ORDER BY ABS(m.mtm_value_inr) DESC
        """)
        rows = [dict(r) for r in cursor.fetchall()]
        conn.close()

        headers = [
            "Trade ID", "Instrument", "Counterparty", "Date",
            "MTM (Trade Ccy)", "Trade Currency", "MTM (₹)", "Daily ΔP&L (₹)",
            "Discount Factor", "Status"
        ]
        self._style_header(ws, 2, headers)

        for i, r in enumerate(rows):
            row = 3 + i
            self._zebra(ws, row, len(headers), i)
            mtm_inr = r.get("mtm_value_inr", 0)
            row_data = [
                r.get("trade_id"), r.get("instrument_type"), r.get("counterparty_name"),
                r.get("valuation_date"),
                f"{r.get('mtm_value', 0):,.0f}", r.get("trade_ccy"),
                f"₹{mtm_inr:,.0f}", f"₹{r.get('pnl_daily', 0):,.0f}",
                round(r.get("discount_factor", 1), 4),
                "IN THE MONEY" if mtm_inr > 0 else "OUT OF MONEY",
            ]
            for col, val in enumerate(row_data, start=1):
                cell = ws.cell(row, col, val)
            # Colour the MTM cell
            mtm_cell = ws.cell(row, 7)
            if mtm_inr > 0:
                mtm_cell.font = Font(color="006100", bold=True)
            elif mtm_inr < 0:
                mtm_cell.font = Font(color="9C0006", bold=True)

        # Summary at bottom
        total_mtm = sum(r.get("mtm_value_inr", 0) for r in rows)
        total_pnl = sum(r.get("pnl_daily", 0) for r in rows)
        sr = 3 + len(rows) + 1
        ws.cell(sr, 1, "PORTFOLIO TOTAL").font = Font(bold=True, color=NAVY)
        ws.cell(sr, 7, f"₹{total_mtm:,.0f}").font = Font(bold=True, color=NAVY)
        ws.cell(sr, 8, f"₹{total_pnl:,.0f}").font = Font(bold=True, color=NAVY)

        self._autofit_columns(ws)

    # ── SHEET 4: Margin Calls ───────────────────────────────────

    def _sheet_margin(self, wb, margin_calls):
        ws = wb.create_sheet("📉 Margin Calls")
        ws.freeze_panes = "A3"
        self._style_title(ws, 1, 1, "MARGIN CALL REGISTER — INITIAL & VARIATION MARGIN", 9)

        headers = [
            "Call ID", "Trade ID", "Counterparty", "Date", "Type",
            "Amount (CCY)", "Currency", "Amount (₹)", "Direction",
            "Status", "Due Date", "Regulatory Basis"
        ]
        self._style_header(ws, 2, headers)

        status_colours = {"MET": "006100", "PENDING": "9C6500", "DISPUTED": "9C0006", "FAILED": "9C0006"}

        for i, m in enumerate(margin_calls):
            r = 3 + i
            self._zebra(ws, r, len(headers), i)
            row_data = [
                m.get("call_id"), m.get("trade_id") or "Portfolio",
                m.get("counterparty_id"), m.get("call_date"),
                m.get("margin_type"), f"{m.get('call_amount', 0):,.0f}",
                m.get("call_currency"), f"₹{m.get('call_amount_inr', 0):,.0f}",
                m.get("direction"), m.get("status"),
                m.get("due_date"), m.get("regulatory_basis"),
            ]
            for col, val in enumerate(row_data, start=1):
                ws.cell(r, col, val)
            status = m.get("status", "")
            ws.cell(r, 10).font = Font(color=status_colours.get(status, "000000"), bold=True)

        # Summary
        sr = 3 + len(margin_calls) + 1
        total_im = sum(m["call_amount_inr"] for m in margin_calls if m["margin_type"] == "INITIAL_MARGIN")
        total_vm = sum(m["call_amount_inr"] for m in margin_calls if m["margin_type"] == "VARIATION_MARGIN")
        ws.cell(sr, 1, f"Total IM Calls: {sum(1 for m in margin_calls if m['margin_type']=='INITIAL_MARGIN')}  |  ").font = Font(bold=True)
        ws.cell(sr, 8, f"₹{total_im/1e7:.1f}Cr IM  +  ₹{total_vm/1e7:.1f}Cr VM").font = Font(bold=True, color=NAVY)
        self._autofit_columns(ws)

    # ── SHEET 5: Collateral ─────────────────────────────────────

    def _sheet_collateral(self, wb, col_summary):
        ws = wb.create_sheet("🏦 Collateral")
        self._style_title(ws, 1, 1, "COLLATERAL MANAGEMENT SUMMARY (RBI + ISDA Eligible)", 6)

        headers = ["Direction", "Collateral Type", "# Postings", "Gross Value (₹)", "Net Value after Haircut (₹)"]
        self._style_header(ws, 2, headers)

        for i, r in enumerate(col_summary):
            row = 3 + i
            self._zebra(ws, row, 5, i)
            ws.cell(row, 1, r.get("direction"))
            ws.cell(row, 2, r.get("collateral_type"))
            ws.cell(row, 3, r.get("count"))
            ws.cell(row, 4, f"₹{r.get('total_gross', 0):,.0f}")
            ws.cell(row, 5, f"₹{r.get('total_net_inr', 0):,.0f}")
            if r.get("direction") == "POSTED":
                ws.cell(row, 1).font = Font(color="9C0006", bold=True)
            else:
                ws.cell(row, 1).font = Font(color="006100", bold=True)

        self._autofit_columns(ws)

    # ── SHEET 6: Reconciliation ─────────────────────────────────

    def _sheet_reconciliation(self, wb, recon_summary):
        ws = wb.create_sheet("✅ Reconciliation")
        self._style_title(ws, 1, 1, "DAILY POSITION RECONCILIATION REPORT", 5)

        total = recon_summary.get("total", 1)
        matched = recon_summary.get("MATCHED", 0)
        auto_res = recon_summary.get("AUTO_RESOLVED", 0)
        breaks = recon_summary.get("BREAK", 0)
        accuracy = recon_summary.get("accuracy_pct", 0)

        data = [
            ("Total Records Reconciled", total),
            ("Matched",                  matched),
            ("Auto-Resolved (small diff)", auto_res),
            ("Open Breaks",              breaks),
            ("Accuracy (%)",             f"{accuracy:.2f}%"),
            ("Target Accuracy",          "99.80%"),
            ("Target Met?",              "✅ YES" if accuracy >= 99.8 else "❌ NO"),
            ("Regulatory Compliant?",    "✅ YES (EMIR >99%)" if accuracy >= 99.0 else "❌ NO"),
        ]

        self._style_header(ws, 2, ["Metric", "Value"])
        for i, (label, value) in enumerate(data):
            r = 3 + i
            self._zebra(ws, r, 2, i)
            ws.cell(r, 1, label).font = Font(bold=(i in (4, 6, 7)))
            cell = ws.cell(r, 2, value)
            if i == 4:
                cell.font = Font(bold=True, size=14, color=NAVY)
            if i == 6:
                cell.font = Font(bold=True, color="006100" if accuracy >= 99.8 else "9C0006")

        # Break categories explanation
        br = 3 + len(data) + 2
        ws.cell(br, 1, "BREAK CATEGORIES & DESCRIPTIONS").font = Font(bold=True, color=NAVY)
        self._style_header(ws, br + 1, ["Category", "Description", "Typical Cause"])
        cats = [
            ("MTM_MISMATCH",      "MTM values differ between parties", "Different pricing models or curves"),
            ("NOTIONAL_MISMATCH", "Notional amounts disagree",         "Booking error or amendment not confirmed"),
            ("DATE_MISMATCH",     "Maturity/start dates differ",       "Different business day conventions"),
            ("MISSING_TRADE",     "Trade on one side only",            "Trade not confirmed by counterparty"),
            ("STATUS_MISMATCH",   "Different lifecycle status",        "Settlement or termination not reflected"),
        ]
        for j, (cat, desc, cause) in enumerate(cats):
            r = br + 2 + j
            self._zebra(ws, r, 3, j)
            ws.cell(r, 1, cat).font = Font(bold=True)
            ws.cell(r, 2, desc)
            ws.cell(r, 3, cause)

        self._autofit_columns(ws)
